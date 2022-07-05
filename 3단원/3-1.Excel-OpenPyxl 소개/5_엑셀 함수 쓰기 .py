#엑셀 함수 사용해보기

#1.Cell에 엑셀 함수 써보기

# import openpyxl as op

# wb=op.load_workbook(r"test.xlsx")
# ws=wb.active

# ws["E11"].value="=SUM(C:C)"

# wb.save("result.xlsx")

#여러 줄에 함수 한번에 적용하기
# import openpyxl as op

# wb=op.load_workbook(r"test.xlsx")
# ws=wb.active()

# #해당 시트의 최대 행값,열값 구하기
# col_max=ws.max_column
# row_max=ws.max_row

# print("최대행값 : ",row_max) 
# print("최대열값 : ",col_max)
   

# import openpyxl as op

# wb=op.load_workbook(r"test.xlsx")
# ws=wb.active

# row_max=ws.max_row

# for row in range(2,row_max+1):
#     ws["E"+str(row)].value="=C"+str(row)+"*"+"D"+str(row)
#     정수 타입의 row를 str()함수를 통해 문자열로 변경함으로써 에러를 해결한다
# wb.save(r"result.xlsx")

import openpyxl as op

wb=op.load_workbook(r"result.xlsx")
ws=wb.active

data=[]

for row in ws.rows:
    data.append(row[4].value)    #결과값이 아닌 엑셀에 적용했던 수식이 저장됨

print(data)


#해결방안:load_workbook()함수의 내부 옵션값을 수정한다
#내부 옵션값: data_only=True(수식이 계산된 값을 읽어온다), read_only=True(읽기 전용으로 읽어온다)

wb=op.load_workbook(r"result.xlsx",data_only=True)

ws=wb.active

data=[]

for row in ws.rows:
    data.append(row[4].value)
    
print(data)

#None이 출력될 경우 불러온 엑셀파일을 한 번 실행시켜서 저장한 후 다시 코드를 실행한다
#엑셀 파일을 직접 실행해야 코드를 통해 입력했던 수식이 계산되고, 계산된 값을 인식하게 된다

#엑셀파일을 실행시키고 저장하는 코드
import win32com.client

#excel 사용할 수 있게 설정
excel=win32com.client.Dispatch("Excel.Application")

#임시 Workbook 객체 생성 및 엑셀 열기
temp_wb=excel.Workbooks.Open(r"result.xlsx")

#저장
temp_wb.Save()

#excel 종료
excel.quit()



#총정리

import  openpyxl  as  op  #openpyxl 모듈 import
import  win32com.client



def  writeFunc():
    wb = op.load_workbook(r"test.xlsx") #Workbook 객체 생성
    ws = wb.active  #활성화 된 시트 객체 생성

    row_max = ws.max_row  #최대 행값 저장

    #for문 통해 2행~최대행까지 반복문
    #range(a,b+1) : a부터 b까지 반복하는 range 구문
    for  row  in  range(2, row_max+1):
    #함수 자동 작성
        ws["E"+str(row)].value = "=C"+str(row)+"*"+"D"+str(row)

    wb.save(r"result.xlsx") #결과 엑셀파일 저장


def  loadData():
    #엑셀 직접 실행시키고 저장 및 종료
    excel = win32com.client.Dispatch("Excel.Application")
    temp_wb = excel.Workbooks.Open(r"result.xlsx")
    temp_wb.Save()
    excel.quit()

    #openpyxl 통해 Workbook 객체 및 WorkSheet 객체 생성
    wb = op.load_workbook(r"result.xlsx", data_only=True)
    ws = wb.active

    #빈 리스트 생성
    data = []

    #ws.rows 속성 활용하여 for문 진행
    for  row  in  ws.rows:
        data.append(row[4].value) #E열 데이터를 리스트에 추가

    print(data) #최종 리스트 출력

if  __name__ == "__main__":
    writeFunc() #writeFunc 실행
    loadData() #loadData 실행
