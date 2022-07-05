#엑셀 서식 지정하기-조건부 서식

import openpyxl as op
from openpyxl.styles.fonts import Font

#test 엑셀 파일이 있는 경로 지정
path=r"C:\Users\신관우\OneDrive\바탕 화면\문서자동화"

wb=op.load_workbook(path+'/'+"조건부서식 test.xlsx",data_only=True)
ws=wb.active

#합/불 판정해주는 함수
def passfail():
    #최대 행값 구하기
    max_row=ws.max_row
    #최대 행값 활용하여 for문(2행부터 ~ 최대행까지)
    for row_index in range(2,max_row+1):
        #평균값 데이터를 average 변수에 저장
        average=ws.cell(row=row_index,column=5).value
        #평균이 70점 이상이면 '합격' 표시
        if average>=70:
            ws.cell(row=row_index,column=6).value="합격" 
        #평균이 70점 미만이면 '불합격' 표시
        else:
            ws.cell(row=row_index,column=6).value="불합격"
        
#합격/불합격에 대한 조건부 서식 적용하기

def condiitonFormat():
    #합격일 때 format 변수로 설정
    pass_format=Font(size=12,name='굴림',color='000000FF')  #000000FF는 파란색
    #불합격일 때 format을 변수로 설정
    fail_format=Font(size=12,name='굴림',color='00FF0000')  #00FF0000은 빨간색
    #행 최대값 구하기
    max_row=ws.max_row
    
    #행 최대값 사용하여 for문 사용(반복)
    for row_index in range(2,max_row+1):
        #합격/불합격인지 문자열 읽어오기
        result_str=ws.cell(row=row_index,column=6).value
        #합격일 경우 서식
        if result_str=="합격":
            ws.cell(row=row_index,column=6).font=pass_format
        #불합격일 경우 서식
        else:
            ws.cell(row=row_index,column=6).font=fail_format
            
#실행부
if __name__=="__main__":
    passfail()
    wb.save(r"result.xlsx")   
    
