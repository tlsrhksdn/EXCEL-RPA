# 데이터가 새로 입력되었을 때 영역을 자동으로 설정해준다.

#1.rows 출력하기

# import openpyxl as op

# wb=op.load_workbook(r"test.xlsx")
# ws=wb["업"]

# print("#rows 출력")
# for row_rng in ws.rows:
#     print(row_rng)  #각 행에 대한 1차원 배열 출력
    
#2.columns 출력하기
import openpyxl as op

wb=op.load_workbook(r"test.xlsx")
ws=wb["업"]

print("#columns 출력")
for col_rng in ws.columns:
    print(col_rng)  #각 행에 대한 1차원 배열 출력
    
#데이터 추가 후 rows 출력해보기

import openpyxl as op

wb=op.load_workbook(r"test.xlsx")
ws=wb["업"]

print("#rows 출력")
for row_rng in ws.rows:
    
    for cell in  row_rng: 
        if cell.value != None:  #None(값이 존재하지 않을)경우 출력 제외
            print(cell.value)
        
    