#openpyxl 설치 및 import하기

#openpyxl 설치

#import하기 

#Workbook 객체 생성하기

#Workbook을 생성하는 경우 2가지 

#새로운 엑셀 파일을 만드는 경우

#openpyxl 모듈 import
import openpyxl as op

#새로운 Workbook 객체 생성
wb=op.Workbook()

#객체 출력
print(wb)

#생성한 객체를 실제 파일로 저장한다
wb.save("openpyxl_test.xlsx")

#경로 안의 r표시 : unicode 에러 발생시 사용
#파이썬에서는 역슬래시를 인식하지 못해 디코더 에러가 발생한다
#경로를 표현할 때 슬래시를 사용하거나, 역 슬래시를 사용해야 할 경우 문자열 앞에 r을 표기하여 준다.

wb.save(r"C:\Users\Desktop\VS CODE\openpyxl_test.xlsx")

#기존에 만들어져 있는 파일로 객체 생성

#load.workbook()함수를 사용해 기존 엑셀 파일을 Workbook 객체로 생성한다

#기존 엑셀 파일을 Workbook 객체로 생성하고 싶다면 .load_workbook() 함수를 사용한다

#파일의 경로를 변수를 통해 지정한다.
#path라는 변수명을 가진 문자열이며,unicode 에러를 방지하기 위해 문자열 앞에 "r"을 표기한다

# path=r"경로"
# wb=op.load_workbook(path+"/test.xlsx")

#WorkSheet 설정하기

#엑셀 파일 내부의 어떤 Sheet에 접근할 것인지 Sheet를 설정한다

#Sheet를 설정하는 방법은 Sheet를 새로 생성하여 객체를 설정하거나 기존 만들어진 Sheet로 설정하는 방법이 있다.

#새로운 Sheet 만들기

#ㅇSheet를 생성할 때 주의해야할 점
#해당 엑셀 파일의 객체 변수를 통해 접근해야 한다.

#openpyxl 모듈 import
import openpyxl as op

#새로운 워크북 객체 생성
wb=op.Workbook()

#wb 객체를 통해 새로운 시트 생성(시트명:new_sheet1)
ws=wb.create_sheet("new_sheet1")

#ws객체 출력해보기 
print(ws)

#해당 워크북(엑셀파일) 저장하기
wb.save("test_result.xlsx")

#기존 만들어져 있는 Sheet에 접근하기

#활성화되어 있는 Sheet에 접근

#활성화되어 있는 Sheet의 경우 .active 코드로 Sheet를 설정할 수 있다.

#활성화되어있는 시트 설정
ws=wb.active

#wb["시트명"]으로 Sheet를 설정할 수 있다

#Workbook 객체의 모든 Sheet명 출력

#.sheetnames()라는 함수를 통해 엑셀 파일의 내부 Sheet목록을 전부 출력할 수 잇다.

#내부 Sheet를 파이썬의 자료 형 중 리스트로 저장하기 때문에 for문과 같이 반복문을 통해 접근하는 게 가능하다.

import openpyxl as op

wb=op.load_workbook(r"경로")
#해당 Workbook의 시트 목록을 리스트로 저장
ws_list=wb.sheetnames

print(ws_list)

for sht in ws_list:
  ws=wb[sht] #Sheet 객체 생성
  print(ws)
  